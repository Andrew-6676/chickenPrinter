import os
import subprocess
import re
import winsound

import openpyxl
from comtypes import CoInitializeEx

from openpyxl.utils.cell import get_column_letter
import win32print

from general.xls import EXCEL
from printer.bar_code import insertBarCode
# ----------------------------------------------------------------------------------------------- #

def beep():
	winsound.Beep(2500, 100)
	winsound.Beep(2500, 100)
	winsound.Beep(2500, 100)

# ----------------------------------------------------------------------------------------------- #
async def genereate_file_to_print(template_file, data):
	"""
	Подстановка значений из data в шаблон и конвертирование его в PDF
	:param data: данные для вставки в шаблон
	:return:
	"""
	current_work_dir = os.getcwd() + '/printer'

	wb = openpyxl.load_workbook(template_file)
	sheet = wb.active
	amountOfRows = sheet.max_row
	amountOfColumns = sheet.max_column

	# проходимся по всем ячейкам, заменяем шаблон на значения
	for i in range(amountOfColumns):
		for k in range(amountOfRows):
			cell = str(sheet[get_column_letter(i + 1) + str(k + 1)].value)
			if cell != 'None':
				newCell = cell
				for v in data:
					newCell = re.sub(r'\{\{' + v + '\}\}', str(data[v]), newCell)
				sheet[get_column_letter(i + 1) + str(k + 1)] = newCell

				# ищем штрихкод
				bar_res = re.findall(r'\{\{(.+)=(ean13|code128),?(\d+)?,?([\d.]+)?\}\}', newCell)
				if bar_res:
					sheet[get_column_letter(i + 1) + str(k + 1)] = ''
					code_data, code_type, rotate, resize = bar_res[0]
					code_data = data[code_data]
					rotate = rotate if rotate else 0
					resize = resize if resize else 1
					# print('-->', get_column_letter(i + 1) + str(k + 1), code_type, code_data, rotate)
					insertBarCode(sheet, get_column_letter(i + 1) + str(k + 1), code_type, code_data, int(rotate), float(resize))

	tmp_file = 'print_total.xlsx' if re.match(r'.*_total.*', template_file) else 'print.xlsx'
	wb.save(current_work_dir + '/tmp/' + tmp_file)

	return current_work_dir + '/tmp/' + tmp_file
# ----------------------------------------------------------------------------------------------- #
def xlsx_to_pdf(src_file):
	"""
	конвертируем XLSX в PDF
	:param src_file: путь к конвертируемому файлу
	:return: путь к получившемуся pdf-файлу
	"""
	current_work_dir = os.getcwd() + '/printer'
	wb_path = src_file
	wb = EXCEL.Workbooks.Open(wb_path)

	ws_index_list = [1]  # say you want to print these sheets

	tmp_file = 'print_total' if re.match(r'.*_total.*', src_file) else 'print'
	path_to_pdf = current_work_dir + '/tmp/' + tmp_file +'.pdf'

	wb.WorkSheets(ws_index_list).Select()
	# ExportAsFixedFormat(type, filename, quality=0-good,1-bad, IncludeDocProperties, IgnorePrintAreas, From, To, OpenAfterPublish)
	wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf, 0, False, False, 1, 1)

	wb.Close()

	return path_to_pdf
	# excel.Quit()

# ----------------------------------------------------------------------------------------------- #
def print_file(pdf_file_name, gs='gswin32c'):
	"""
	Выводим на принтер pdf-файл
	:param pdf_file_name:
	:return:
	"""
	# GHOSTSCRIPT_PATH = "w:\\#### D\Work\\# python\\tsc_sample\\gsprint\\gsprint.exe"
	# GSPRINT_PATH = "C:\\Program Files (x86)\\gs\\gs9.50\\bin\\gswin32.exe"
	#
	printer_name = win32print.GetDefaultPrinter()
	print(f'printer - [{printer_name}]')

	#
	# #win32api.ShellExecute (0, "print", filename, f'"{win32print.GetDefaultPrinter()}"', ".", 0)
	# win32api.ShellExecute(0, 'open', GSPRINT_PATH, '-ghostscript "'+GHOSTSCRIPT_PATH+'" -printer "'+printer_name+'" "1.pdf"', '.', 0)

	args = gs + '.exe ' \
	       '-sDEVICE=mswinpr2 ' \
	       '-q ' \
	       '-dNOPROMPT -dFitPage ' \
	       '-dBATCH ' \
	       '-dNOPAUSE ' \
	       '-sOutputFile="%printer%{}" '.format(printer_name)
	ghostscript = args + f'"{pdf_file_name}"'
	subprocess.call(ghostscript, shell=True)
