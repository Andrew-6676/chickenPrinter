import falcon
import json
from colorama import Fore, Style

from config import Config
from report import Report
from openpyxl import Workbook

#--------------------------------------------------------------------------------------------------#
def log(req, resp, resource, params):
	print((Fore.RED + '{}: ' + Fore.GREEN + '{}' + Style.RESET_ALL).format(req.method, req.path))

def set_json(req, resp, resource, params):
	resp.content_type = 'application/json'
#--------------------------------------------------------------------------------------------------#
@falcon.before(log)
class Index():
	def __init__(self, q=None, config=None):
		self.qqq=q
		self.config=config

	def on_get(self, req, resp, action=None):
		# resp.body = filename
		if not action:
			resp.content_type = 'text/html'
			with open('./static/index.html', 'r') as f:
				print('index')
				resp.body = f.read()

#--------------------------------------------------------------------------------------------------#
@falcon.before(log)
@falcon.before(set_json)
class Production():
	def __init__(self, db_connection, q=None, config=None):
		self.qqq=q
		self.config = config
		self.conn = db_connection
		self.cursor = db_connection.cursor()

	def on_get(self, req, resp, id=None):
		if id:
			sql = """select * from production where id={} and deleted=0""".format(id)
			self.cursor.execute(sql)
			data = self.cursor.fetchone()
		else:
			sql = """select * from production where not deleted"""
			self.cursor.execute(sql)
			data = self.cursor.fetchall()

		resp.body = json.dumps(data)

	def on_post(self, req, resp):
		prod = json.load(req.stream)
		print('post ', prod)
		sql = "insert into production " \
		      "(group_name, name, descr, ingridients, storage_conditions, " \
		      "nutritional_value, energy_value, RC_BY, TU_BY, STB, " \
		      "expiration_date, bar_code) " \
		      "values" \
		      f"('{prod['group_name']}', '{prod['name']}', '{prod['descr']}', '{prod['ingridients']}', '{prod['storage_conditions']}', " \
		      f"'{prod['nutritional_value']}', '{prod['energy_value']}', '{prod['RC_BY']}', '{prod['TU_BY']}', '{prod['STB']}', " \
		      f"'{prod['expiration_date']}', '{prod['bar_code']}')"
		try:
			self.cursor.execute(sql)
			self.conn.commit()
			resp.body = json.dumps({'status': 'ok'})
		except Exception as e:
			self.conn.rollback()
			print(e)
			resp.body = json.dumps({'status': 'error'})

	def on_put(self, req, resp, id=None):
		prod = json.load(req.stream)
		print('put ', prod)
		sql = "update production " \
			  "set " \
		      f"group_name='{prod['group_name']}', " \
		      f"name='{prod['name']}', " \
		      f"descr='{prod['descr']}', " \
		      f"ingridients='{prod['ingridients']}', " \
		      f"storage_conditions='{prod['storage_conditions']}', " \
		      f"nutritional_value='{prod['nutritional_value']}', " \
		      f"energy_value='{prod['energy_value']}', " \
		      f"RC_BY='{prod['RC_BY']}', " \
		      f"TU_BY='{prod['TU_BY']}', " \
		      f"STB='{prod['STB']}', " \
		      f"expiration_date='{prod['expiration_date']}', " \
		      f"bar_code='{prod['bar_code']}' where id={id}"
		try:
			self.cursor.execute(sql)
			self.conn.commit()
			resp.body = json.dumps({'status': 'ok'})
		except Exception as e:
			self.conn.rollback()
			print(e)
			resp.body = json.dumps({'status': 'error'})

	def on_delete(self, req, resp, id=None):
		print('delete prod', id)
		sql = f"update production set deleted=1 where id={id}"
		try:
			self.cursor.execute(sql)
			self.conn.commit()
			resp.body = json.dumps({'status': 'ok'})
		except Exception as e:
			self.conn.rollback()
			print(e)
			resp.body = json.dumps({'status': 'error'})
#--------------------------------------------------------------------------------------------------#
@falcon.before(log)
@falcon.before(set_json)
class User():
	def __init__(self, db_connection, q=None, config=None):
		self.qqq=q
		self.config = config
		self.conn = db_connection
		self.cursor = db_connection.cursor()

	def on_get(self, req, resp, id=None):
		if id:
			sql = """select * from user where id={} and deleted=0""".format(id)
			self.cursor.execute(sql)
			data = self.cursor.fetchone()
		else:
			sql = """select * from user where not deleted"""
			self.cursor.execute(sql)
			data = self.cursor.fetchall()


		resp.body = json.dumps(data)

	def on_post(self, req, resp):
		user = json.load(req.stream)
		print('post ', user)
		sql = f"insert into user (name, pass) values('{user['name']}', '{user['pass']}')"
		try:
			self.cursor.execute(sql)
			self.conn.commit()
			resp.body = json.dumps({'status': 'ok'})
		except Exception as e:
			self.conn.rollback()
			print(e)
			resp.body = json.dumps({'status': 'error'})

	def on_put(self, req, resp, id=None):
		user = json.load(req.stream)
		print('put ', id, user)
		sql = f"update user set name='{user['name']}', pass='{user['pass']}' where id={id}"
		try:
			self.cursor.execute(sql)
			self.conn.commit()
			resp.body = json.dumps({'status': 'ok'})
		except Exception as e:
			self.conn.rollback()
			print(e)
			resp.body = json.dumps({'status': 'error'})

	def on_delete(self, req, resp, id=None):
		print('del ', id)
		sql = f'update user set deleted=1 where id={id}'
		try:
			self.cursor.execute(sql)
			self.conn.commit()
			resp.body = json.dumps({'status': 'ok'})
		except Exception as e:
			self.conn.rollback()
			print(e)
			resp.body = json.dumps({'status': 'error'})

#--------------------------------------------------------------------------------------------------#

# class Reporter(Report):
# 	def __init__(self, database, q=None, config=None):
# 		self.qqq=q
# 		self.config = config
# 		Report.__init__(self, database, config)
#
# 	def on_options(self, req, res):
# 		res.set_header('Access-Control-Allow-Origin', '*')
# 		res.set_header('Access-Control-Allow-Methods', 'GET,POST,PUT')
# 		res.set_header('Access-Control-Allow-Headers', 'Content-Type')
#
# 	def on_get(self, req, resp, action='data'):
# 		resp.set_header('Access-Control-Allow-Origin', '*')
# 		resp.set_header('Access-Control-Allow-Methods', 'GET')
# 		resp.set_header('Access-Control-Allow-Headers', 'Content-Type')
#
# 		res = None
# 		if action=='data':
# 			#res = {'data': Report.getReport(self, req.params, )}
# 			res = {
# 				'partNumber': self.qqq.getPartNumber(),
# 				'partDate': self.qqq.getPartDate(),
# 				'data': self.qqq.getRows(),
# 				'dataErr': self.qqq.getRowsErr()
# 			}
# 		else:
# 			print(Fore.BLACK + Style.BRIGHT + 'GET: report('+action+')' + Style.RESET_ALL)
# 		# if action == 'getOnePart':
# 		# 	res = Report.getOnePart(self, req.params.get('date'), req.params.get('num'))
# 		if action == 'getAllParts':
# 			res = Report.getLastParts(self, req.params.get('limit'))
# 		if action == 'getLastParts':
# 			res = Report.getLastParts(self, req.params.get('limit', 10))
# 		if action=='getPartData':
# 			res = Report.getPartData(self, req.params.get('date'), req.params.get('num', 0))
# 		if action=='getDay':
# 			res = Report.getDay(self, req.params.get('date'))
# 		if action=='excel_part':
# 			data = Report.getPartData(self, req.params.get('date'), req.params.get('num', 0))
# 			header = 'Варка № {} от {}'.format(req.params.get('num', 0), req.params.get('date'))
# 			res = {'link': self.makeExcel(data,  header=header, sum=[1], title=('Время', 'Вес'))}
# 		if action=='excel_day':
# 			data = Report.getDay(self, req.params.get('date'))
# 			header = 'Отчёт за {}'.format(req.params.get('date'))
# 			res = {'link': self.makeExcel(data, header=header, sum=[1,2], title=('Номер варки', 'Вес', 'Кол-во голов', 'Мин. вес', 'Средний вес', 'Макс. вес'))}
# 		if action=='stat':
# 			res = Report.getStat(self)
# 			# print(self.qqq.getRows())
# 		if action=='pack':
# 			print(Fore.BLACK + Style.BRIGHT + 'POST: report(' + action + ')' + Style.RESET_ALL)
# 			res = Report.packData(self, req.params)
#
# 		resp.body = json.dumps(res)
#
# 	def makeExcel(self, data, header='Заголовок отчёта', sum=[1], title=('Время','Вес')):
# 		wb = Workbook()
# 		ws = wb.active
#
# 		ws['A1'] = header
#
# 		i = 0
# 		for t in title:
# 			ws['ABCDEFG'[i] + str(3)] = t
# 			i += 1
#
# 		r = 4
# 		_sum = {}
# 		for i in sum:
# 			_sum[i] = 0
#
# 		for row in data:
# 			# ws['A'+str(r)] = row['time'] #time.strftime('%d.%m.%Y', time.strptime(row[1], "%Y-%m-%d"))
# 			# ws['B'+str(r)] = row['weight'] #round(row[2])
# 			for i in sum:
# 				_sum[i] += row[i]
#
# 			i = 0
# 			for f in row:
# 				ws['ABCDEFG'[i] + str(r)] = f
# 				i += 1
#
# 			r += 1
# 		ws['A'+str(r)] = 'ИТОГО:'
# 		# ws['B'+str(r)] = '=СУММ(B1:B'+str(r-1)+')'
# 		for i in sum:
# 			ws['ABCDEFG'[i]+str(r)] = _sum[i]
#
# 		file = './static/reports/report.xlsx'
# 		wb.save(file)
#
# 		return 'reports/report.xlsx'
#
# 	def on_put(self, req, res):
# 		res.set_header('Access-Control-Allow-Origin', '*')
# 		res.set_header('Access-Control-Allow-Methods', 'PUT')
# 		res.set_header('Access-Control-Allow-Headers', 'Content-Type')
# 		print(Fore.BLACK + Style.BRIGHT + 'Put: report()' + Style.RESET_ALL)
#
# 		data = json.load(req.stream)
# 		sql= """update days_data set total_weight="""+str(data['val'])+""" where id="""+str(data['id'])
# 		print(sql)
#
# 		try:
# 			self.cursor.execute(sql)
# 			self.conn.commit()
# 			res.body = json.dumps({'status': 'ok'})
# 		except Exception as e:
# 			self.conn.rollback()
# 			res.body = json.dumps({'status': 'error'})
		
#--------------------------------------------------------------------------------------------------#
